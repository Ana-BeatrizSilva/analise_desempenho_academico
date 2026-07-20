# Importações
from openpyxl import Workbook
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from pathlib import Path

# Configuração das pastas do projeto

PASTA_RELATORIOS = Path("relatorios")
PASTA_GRAFICOS = Path("graficos")
PASTA_IMAGENS = Path("imagens")

PASTA_RELATORIOS.mkdir(exist_ok=True)
PASTA_GRAFICOS.mkdir(exist_ok=True)
PASTA_IMAGENS.mkdir(exist_ok=True)

# Funções auxiliares
def obter_situacao(media):
    if media >= 7:
        return "Aprovação"
    elif media >= 5:
        return "Recuperação"
    return "Reprovação"

def solicitar_nota(numero_nota):
    while True:
        try:
            nota = float(input(f"Informe a nota {numero_nota}: "))

            if 0 <= nota <= 10:
                return nota

            print("A nota deve estar entre 0 e 10.")

        except ValueError:
            print("Valor inválido. Digite um número.")

def configurar_planilha():
    planilha = Workbook()
    aba = planilha.active
    aba.title = "Alunos"
    aba.freeze_panes = "A2"

    aba.append([
        "Nome",
        "Nota 1",
        "Nota 2",
        "Nota 3",
        "Nota 4",
        "Média",
        "Situação"
    ])

    aba.auto_filter.ref = "A1:G1"

    for celula in aba[1]:
        celula.font = Font(
            bold=True
        )

    aba.column_dimensions["A"].width = 25
    aba.column_dimensions["B"].width = 10
    aba.column_dimensions["C"].width = 10
    aba.column_dimensions["D"].width = 10
    aba.column_dimensions["E"].width = 10
    aba.column_dimensions["F"].width = 12
    aba.column_dimensions["G"].width = 15

    return planilha, aba

def cadastrar_alunos(aba):
    medias = []
    nomes = []
    alunos = []

    aprovacao = 0
    recuperacao = 0
    reprovacao = 0

    qtd_alunos = int(input("Quantidade de alunos que deseja cadastrar: "))

    for i in range(qtd_alunos):
        nome = input(f"Informe o nome do {i+1}° aluno: ")

        notas = []

        for j in range(4):
            nota = solicitar_nota(j + 1)
            notas.append(nota)

        media = sum(notas) / len(notas)

        situacao = obter_situacao(media)

        if situacao == "Aprovação":
            aprovacao += 1
        elif situacao == "Recuperação":
            recuperacao += 1
        else:
            reprovacao += 1

        nomes.append(nome)
        medias.append(media)

        alunos.append({
            "nome": nome,
            "media": media,
            "situacao": situacao
        })

        aba.append([
            nome,
            notas[0],
            notas[1],
            notas[2],
            notas[3],
            round(media, 2),
            situacao
        ])

    return medias, nomes, alunos, aprovacao, recuperacao, reprovacao

def criar_aba_estatisticas(
    planilha,
    medias,
    aprovacao,
    recuperacao,
    reprovacao
):
    media_turma = sum(medias) / len(medias)
    maior_media = max(medias)
    menor_media = min(medias)

    aba_estatisticas = planilha.create_sheet("Estatísticas")

    aba_estatisticas.append([
        "Indicador",
        "Valor"
    ])

    aba_estatisticas.append([
        "Média da turma",
        round(media_turma, 2)
    ])

    aba_estatisticas.append([
        "Maior média",
        round(maior_media, 2)
    ])

    aba_estatisticas.append([
        "Menor média",
        round(menor_media, 2)
    ])

    aba_estatisticas.append([
        "Aprovados",
        aprovacao
    ])

    aba_estatisticas.append([
        "Recuperação",
        recuperacao
    ])

    aba_estatisticas.append([
        "Reprovados",
        reprovacao
    ])

    for celula in aba_estatisticas[1]:
        celula.font = Font(bold=True)

    return media_turma, maior_media, menor_media

def criar_ranking(planilha, alunos):

    alunos_ordenados = sorted(
        alunos,
        key=lambda aluno: aluno["media"],
        reverse=True
    )

    aba_ranking = planilha.create_sheet("Ranking")

    aba_ranking.append([
        "Posição",
        "Nome",
        "Média",
        "Situação"
    ])

    for posicao, aluno in enumerate(alunos_ordenados, start=1):
        aba_ranking.append([
            posicao,
            aluno["nome"],
            round(aluno["media"], 2),
            aluno["situacao"]
        ])

    for celula in aba_ranking[1]:
        celula.font = Font(bold=True)

def exibir_estatisticas(
    media_turma,
    maior_media,
    menor_media,
    aprovacao,
    recuperacao,
    reprovacao
):
    total_alunos = aprovacao + recuperacao + reprovacao

    percentual_aprovacao = (aprovacao / total_alunos) * 100
    percentual_recuperacao = (recuperacao / total_alunos) * 100
    percentual_reprovacao = (reprovacao / total_alunos) * 100

    print("\n=== ESTATÍSTICAS DA TURMA ===")

    print(f"Quantidade de alunos: {total_alunos}")

    print(f"\nMédia da turma: {media_turma:.2f}")
    print(f"Maior média: {maior_media:.2f}")
    print(f"Menor média: {menor_media:.2f}")

    print(
        f"\nAprovados: {aprovacao} "
        f"({percentual_aprovacao:.1f}%)"
    )

    print(
        f"Recuperação: {recuperacao} "
        f"({percentual_recuperacao:.1f}%)"
    )

    print(
        f"Reprovados: {reprovacao} "
        f"({percentual_reprovacao:.1f}%)"
    )

def salvar_planilha(planilha):
    caminho = PASTA_RELATORIOS / "relatorio_desempenho_academico.xlsx"
    planilha.save(caminho)
    print("\nPLANILHA GERADA COM SUCESSO!")

def gerar_grafico_medias(alunos, media_turma):

    alunos_ordenados = sorted(
        alunos,
        key=lambda aluno: aluno["media"],
        reverse=True
    )

    nomes = [
        aluno["nome"]
        for aluno in alunos_ordenados
    ]

    medias = [
        aluno["media"]
        for aluno in alunos_ordenados
    ]

    plt.figure(figsize=(10, 6))

    barras = plt.bar(
        nomes,
        medias
    )

    plt.axhline(
        media_turma,
        linestyle="--",
        label=f"Média da turma: {media_turma:.2f}"
    )

    for barra, media in zip(barras, medias):
        plt.text(
            barra.get_x() + barra.get_width()/2,
            barra.get_height(),
            f"{media:.1f}",
            ha="center",
            va="bottom"
        )

    plt.title("Desempenho dos alunos")
    plt.xlabel("Aluno")
    plt.ylabel("Média")

    plt.ylim(0, 10)

    plt.xticks(
        rotation=45,
        ha="right"
    )

    plt.legend()

    plt.tight_layout()

    plt.savefig(
        PASTA_GRAFICOS / "grafico_medias_turma.png"
    )

    plt.show()

def gerar_grafico_situacoes(aprovacao, recuperacao, reprovacao):
    situacoes = [aprovacao, recuperacao, reprovacao]
    labels = ["Aprovação", "Recuperação", "Reprovação"]

    plt.figure()
    plt.pie(situacoes, labels=labels, autopct="%1.1f%%")
    plt.title("Situação da Turma")
    plt.savefig(PASTA_GRAFICOS / "grafico_situacoes_turma.png")
    plt.show()

def main():
    # Configuração
    planilha, aba = configurar_planilha()
    # Cadastro
    medias, nomes, alunos, aprovacao, recuperacao, reprovacao = cadastrar_alunos(aba)
    # Estatísticas
    media_turma, maior_media, menor_media = criar_aba_estatisticas(
        planilha,
        medias,
        aprovacao,
        recuperacao,
        reprovacao
    )
    # Ranking
    criar_ranking(
        planilha,
        alunos
    )

    exibir_estatisticas(
        media_turma,
        maior_media,
        menor_media,
        aprovacao,
        recuperacao,
        reprovacao
    )
    # Relatórios
    salvar_planilha(planilha)

    gerar_grafico_medias(
    alunos,
    media_turma
)

    gerar_grafico_situacoes(
        aprovacao,
        recuperacao,
        reprovacao
    )

if __name__ == "__main__":
    main()